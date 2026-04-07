import openpyxl
import json
import sys
from openpyxl.utils import get_column_letter
from datetime import datetime, date, time
import decimal

def datetime_to_excel_serial(dt):
    """
    Convert datetime/date object to Excel serial number (1900 date system).
    
    Excel stores dates as the number of days since January 1, 1900.
    Note: Excel incorrectly treats 1900 as a leap year, so we need to account for that.
    
    Args:
        dt: datetime, date, or time object
        
    Returns:
        float: Excel serial number
    """
    # Excel's epoch: January 1, 1900 (but Excel treats it as January 0, 1900)
    excel_epoch = datetime(1899, 12, 30)
    
    # Handle time objects separately
    if isinstance(dt, time):
        # Time only: fraction of a day
        total_seconds = dt.hour * 3600 + dt.minute * 60 + dt.second + dt.microsecond / 1_000_000
        return total_seconds / 86400  # 86400 seconds in a day
    
    # Handle date objects (convert to datetime for calculation)
    if isinstance(dt, date) and not isinstance(dt, datetime):
        dt = datetime.combine(dt, time.min)
    
    # Calculate the difference in days
    delta = dt - excel_epoch
    excel_serial = delta.total_seconds() / 86400
    
    return excel_serial


def normalize_cell_value(value):
    """
    Normalize cell values for JSON serialization.
    
    - datetime objects: Convert to Excel serial number (double)
    - date objects: Convert to Excel serial number (double)
    - time objects: Convert to Excel serial number (fraction of day)
    - Decimal: Convert to float and round to 15 decimals
    - Float: Round to 15 decimals
    - None: Keep as None (null in JSON)
    - Other types: Convert to string if not JSON serializable
    
    Args:
        value: Cell value from openpyxl
        
    Returns:
        JSON-serializable value
    """
    # Handle None/null values
    if value is None:
        return None
    
    # For datetime/date/time objects, convert to Excel serial number
    if isinstance(value, (datetime, date, time)):
        excel_serial = datetime_to_excel_serial(value)
        return round(excel_serial, 15)
    
    # Handle Decimal types (convert to float and round)
    if isinstance(value, decimal.Decimal):
        return round(float(value), 15)
    
    # Handle numeric types (int, float)
    if isinstance(value, (int, float)):
        # Keep integers as-is
        if isinstance(value, int):
            return value
        
        # Handle float
        if isinstance(value, float):
            # Check for special float values
            if value != value:  # NaN check
                return None
            if value == float('inf') or value == float('-inf'):
                return str(value)
            
            # Round to 15 decimal places
            return round(value, 15)
        
        return value
    
    # Handle boolean
    if isinstance(value, bool):
        return value
    
    # Handle strings
    if isinstance(value, str):
        return value
    
    # For any other type, try to convert to string
    try:
        # Try JSON serialization test
        json.dumps(value)
        return value
    except (TypeError, ValueError):
        return str(value)



def excel_to_json(excel_file_path, keep_excel_datetime_format=False):
    """
    Convert Excel file to nested JSON structure.
    
    Structure:
    {
        "Sheet1": {
            "A": {1: "value", 2: "value", ...},
            "B": {1: "value", 2: "value", ...},
            ...
        },
        "Sheet2": {
            ...
        }
    }
    
    Args:
        excel_file_path (str): Path to the Excel file
        keep_excel_datetime_format (bool): If True, keeps Excel's numeric datetime representation
        
    Returns:
        dict: Nested dictionary with the structure described above
    """
    try:
        # Load the workbook with data_only=True to get calculated values
        workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
        
        # Initialize the output dictionary
        output_json = {}
        
        # Iterate through all sheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Initialize dictionary for this sheet
            sheet_data = {}
            
            # Get the number of columns
            max_col = sheet.max_column
            
            # Initialize dictionaries for each column using column letters
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                sheet_data[col_letter] = {}
            
            # Iterate through all rows
            for row_idx in range(1, sheet.max_row + 1):
                row = sheet[row_idx]
                
                # Add each cell value to corresponding column dictionary with row number as key
                for col_idx, cell in enumerate(row, start=1):
                    col_letter = get_column_letter(col_idx)
                    if col_letter in sheet_data:
                        # Normalize the cell value for JSON serialization
                        normalized_value = normalize_cell_value(cell.value)
                        sheet_data[col_letter][row_idx] = normalized_value
            
            # Add sheet data to output
            output_json[sheet_name] = sheet_data
        
        return output_json
    
    except FileNotFoundError:
        print(f"Error: File '{excel_file_path}' not found.")
        return None
    except Exception as e:
        print(f"Error processing Excel file: {str(e)}")
        return None


def main():
    """Main function to handle command line execution"""
    
   
    # Get input file path
    input_file = "input_files/gaulEngine.xlsx"
    
    # Get output file path (optional)
    output_file = "input_files/expected_output.json"
    
    print(f"Reading Excel file: {input_file}")
    
    # Convert Excel to JSON
    result = excel_to_json(input_file)
    
    if result is not None:
        # Convert to JSON string with pretty formatting
        json_output = json.dumps(result, indent=2, ensure_ascii=False)
        
        # Save to file if output path is provided
        if output_file:
            try:
                with open(output_file, 'w', encoding='utf-8') as f:
                    f.write(json_output)
                print(f"JSON output saved to: {output_file}")
            except Exception as e:
                print(f"Error saving output file: {str(e)}")
        else:
            # Print to console
            print("\nJSON Output:")
            print(json_output)
        
        print(f"\nSuccessfully processed {len(result)} sheet(s)")


if __name__ == "__main__":
    main()